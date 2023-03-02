from django.shortcuts import render
from rest_framework import viewsets
from .serializers import SurveySerializer, EvaluationSerializer, ParticipantEvaluationSerializer
from .models import Survey, Evaluation, ParticipantEvaluation
from openpyxl import Workbook
from rest_framework.decorators import api_view
from django.http import HttpResponse

@api_view(['GET'])
def export_surveys_to_excel(request):
    survey = Survey.objects.all()
    serializer = SurveySerializer(survey, many=True)
    data = serializer.data
    # Export data to Excel using openpyxl
    workbook = Workbook()
    worksheet = workbook.active
    # Write headers
    headers = [
            'client_id',
            'age',
            'gender',
            'time_personal_home_life',
            'good_performance',
            'employment',
            'pleasant_people',
            'work_interesting',
            'consulted_by_superiors',
            'desirable_area',
            'respected_by_family',
            'chances_for_promotion',
            'keeping_time_for_fun',
            'moderation',
            'service_to_friend',
            'thrift',
            'nervous_tense',
            'happiness',    
            'prevention_due_to_circumstances',
            'state_of_health',
            'proud_citizen',
            'contradicting_superior',
            'good_manager',
            'persistent_efforts',
            'organization_structure',
            'organization_rules',
            'education',
            'job',
            'nationality',
            'nationality_from_birth'
            ]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    # Write data
    for row_num, row_data in enumerate(data, 2):
        row = [row_data['client_id'],
                row_data['age'],
                row_data['gender'],
                row_data['time_personal_home_life'],
                row_data['good_performance'], 
                row_data['employment'], 
                row_data['pleasant_people'], 
                row_data['work_interesting'], 
                row_data['consulted_by_superiors'], 
                row_data['desirable_area'],
                row_data['respected_by_family'],
                row_data['chances_for_promotion'],
                row_data['keeping_time_for_fun'],
                row_data['moderation'],
                row_data['service_to_friend'],
                row_data['thrift'],
                row_data['nervous_tense'],
                row_data['happiness'],
                row_data['prevention_due_to_circumstances'],
                row_data['state_of_health'],
                row_data['proud_citizen'],
                row_data['contradicting_superior'],
                row_data['good_manager'],
                row_data['persistent_efforts'],
                row_data['organization_structure'],
                row_data['organization_rules'],
                row_data['education'],
                row_data['job'],
                row_data['nationality'],
                row_data['nationality_from_birth'],
                ]
        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cultural-dimensions.xlsx'
    workbook.save(response)
    return response

class SurveyView(viewsets.ModelViewSet):
    serializer_class = SurveySerializer
    queryset = Survey.objects.all()
    
class EvaluationView(viewsets.ModelViewSet):
    serializer_class = EvaluationSerializer
    queryset = Evaluation.objects.all()
    
class ParticipantEvaluationView(viewsets.ModelViewSet):
    serializer_class = ParticipantEvaluationSerializer
    queryset = ParticipantEvaluation.objects.all()
    
    # Add this CBV
class Assets(View):

    def get(self, _request, filename):
        path = os.path.join(os.path.dirname(__file__), 'static', filename)

        if os.path.isfile(path):
            with open(path, 'rb') as file:
                return HttpResponse(file.read(), content_type='application/javascript')
        else:
            return HttpResponseNotFound()
